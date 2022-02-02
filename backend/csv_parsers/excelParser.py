import openpyxl
from typing import *

from settings import DB


class NumberInSports:
    def __init__(self):
        self.values = {}

    def __getitem__(self, item):
        return self.values[item]

    def __setitem__(self, item, value):
        self.values[item] = value

    def __contains__(self, value):
            return value in self.values

    def save(self):
        for key in self.values.keys():
            DB.importNumberInSports( key, self.values[key])


class MaxPointsInSport:
    def __init__(self):
        self.values = {}

    def __getitem__(self, item):
        return self.values[item]

    def __setitem__(self, item, value):
        self.values[item] = value

    def __contains__(self, value):
        return value in self.values

    def save(self):
        for key in self.values.keys():
            DB.importMaxPointsInSport( key, self.values[key])



class TotalCountryPoints:
    def __init__(self):
        self.values = {}

    def __getitem__(self, item):
        return self.values[item]

    def __setitem__(self, item, value):
        self.values[item] = value

    def __contains__(self, value):
        return value in self.values

    def save(self):
        for key in self.values.keys():
            DB.importTotalCountryPoints( key, self.values[key])


class CountryBestOrder:
    def __init__(self):
        self.values = {}

    def __getitem__(self, item):
        return self.values[item]

    def __setitem__(self, item, value):
        self.values[item] = value

    def __contains__(self, value):
        return value in self.values

    def save(self):
        for key in self.values.keys():
            DB.importCountryBestOrder( key, self.values[key])


class SuccessRecord:
    def __init__(self, rank : int, country_id : int, points:float):
        self.rank = rank
        self.country_id = country_id
        self.points = points

    def __str__(self) -> str:
        return f"{self.country_id}, {self.points}, {self.rank}"


class SportSuccess:

    def __init__(self, sport_id : str):
        self.sport_id = sport_id
        self.records = []

    def set_max_points(self, points : int):
        self.max_points = points

    def set_num_of_states(self, value : int):
        self.max_points = value

    def __str__(self) -> str:
        s = ""
        for r in self.records:
            s += f"{self.sport_id}, {r} \n"

        return s

    def save(self):
        for record in self.records:

            #sport_id: id, country_id: id, points: float, orders: int
            DB.importSuccessdata( self.sport_id, record.country_id, record.points, record.rank )


class InterconnectnessRecord:

    def __init__(self, countryA: str, countryB: str, type: str, value: float):
        self.countryA = countryA
        self.countryB = countryB
        self.type = type
        self.value = round(value, 5)

    def __str__(self) -> str:
        return f"{self.type} {self.countryA} {self.countryB} {self.value}"

    def save(self):
        DB.importInterconnectednessData(self.type, self.countryA, self.countryB, self.value)


class ParseError(Exception):
    ...


class excelParser:

    def __init__(self):
        ...

    def parseSuccess(self , wb ) -> List:

        sports = DB.getAllSports()
        countries = DB.getActiveCountryIds()

        def getCountryID( name ):
            for item in countries:
                if item['name'] == name.strip():
                    return item["id"]

            return -1

        def getSportID( title ):
            for item in sports:
                if item['title'] == title.strip().replace("/", " AND "):
                    return item["code"]
            return -1


        pocet_statov = NumberInSports()
        max_bodov = MaxPointsInSport()
        pocet_bodov = TotalCountryPoints()
        najvyssie_um = CountryBestOrder()

        success = []
        unknown_sports = []

        sport_row_index = 2

        for sheet in wb.worksheets:

            last_row = sheet.max_row  # ako najhlbsie ide
            last_col = sheet.max_column  # ako najdalej

            for sport_col_index in range( 2, last_col+1, 4 ):

                cell = sheet.cell(row=sport_row_index, column=sport_col_index)
                if cell.value != None:
                    if isinstance(cell.value, str):
                        sport_title = cell.value
                        sport_id = getSportID(sport_title)
                        sport_record = SportSuccess(sport_id)

                        pocet_statov[sport_id] = 0
                        max_bodov[sport_id] = 0
                        for r in range(sport_row_index + 2, last_row + 1):

                            points = sheet.cell(row=r, column=sport_col_index + 1).value

                            if points is not None:

                                points = round(float(points),3)
                                rank = int(sheet.cell(row=r, column=sport_col_index - 1).value)
                                country = sheet.cell(row=r, column=sport_col_index).value
                                country_id = getCountryID(country)


                                if points > max_bodov[sport_id]:
                                    max_bodov[sport_id] = points

                                pocet_statov[sport_id] = pocet_statov[sport_id] + 1


                                if country_id > 0:

                                    if country_id not in pocet_bodov:
                                        pocet_bodov[country_id] = points
                                    else:
                                        pocet_bodov[country_id] = pocet_bodov[country_id] + points


                                    if country_id not in najvyssie_um:
                                        najvyssie_um[country_id] = rank
                                    else:
                                        if rank < najvyssie_um[country_id]:
                                            najvyssie_um[country_id] = rank

                                    sport_record.records.append(SuccessRecord(rank, country_id, points))

                if sport_id > 0:
                    success.append(sport_record)
                else:
                    unknown_sports.append(sport_title)



        return [ success, unknown_sports, pocet_statov, max_bodov, pocet_bodov, najvyssie_um ]

    def parseInterconnectness(self, wb, type: int) -> List[InterconnectnessRecord]:

        sheet = wb.active
        countries = DB.getActiveCountryTranslations()

        def getCountryID(name):
            for item in countries:
                if item['translation'] == name.strip():
                    return item["id"]
            return -1

        LAST_ROW = sheet.max_row  # kolko rows obsahuje zdroj
        country_index = 1
        last_country_index = 2

        for row in range(2, LAST_ROW + 1):

            if sheet.cell(row=row, column=country_index).value is None:
                break

            last_country_index = row
            country1 = sheet.cell(row=row, column=country_index).value  # country(x,1)
            country2 = sheet.cell(row=country_index, column=row).value  # country(1,x)

            if country1 != country2:
                raise ParseError("Inconsistance in countries in first row and first col.")

            if getCountryID(country1) == -1:
                raise ParseError(f"Unknown country in row : {country1}")

            if getCountryID(country2) == -1:
                raise ParseError(f"Unknown country in column : {country2}")

        records = []
        for row in range(2, last_country_index + 1):
            sucet = 0

            c1 = sheet.cell(row=row, column=1)
            country1 = getCountryID(c1.value.strip())

            for col in range(2, last_country_index + 1):

                if row != col:

                    cell = sheet.cell(row=row, column=col)
                    value = float(cell.value)

                    if value != 0:
                        c2 = sheet.cell(row=1, column=col)
                        country2 = getCountryID(c2.value.strip())

                        records.append(InterconnectnessRecord(country1, country2, type, value))

                        sucet += value

            if abs(1 - sucet) > 0.00000000000005:  # priemerna odchylka = 2.62662520460098e-16
                raise ParseError(f'Sum of interconnectedness for country {c1.value} is not equal to 1')

        return records

# p = excelParser()

# example of usage - Success
# wb = openpyxl.load_workbook(filename='ALL SPORTS RANKING 2019.xlsx')
# parsed = p.parseSuccess(wb)

# s= 0
# nieco = []

# for r in parsed[0]:
#     s += len(r.records)
#     nieco.append( [r.sport_id , len(r.records)] )

# for n in nieco:
#     print ( n )


# example of usage - Interconnectness

# parsed = p.parseInterconnectness("intercon_test.xlsx", 1)
# print(len(parsed), "velkost")
